from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.db.base import JobStatus
from app.db.models import Fund
from app.imports.tasks import process_next_job

from .test_import_publish_dashboard import _valuation_xlsx


def _history_workbook(day: str, nav: str) -> bytes:
    workbook = load_workbook(BytesIO(_valuation_xlsx()))
    sheet = workbook.active
    sheet.cell(3, 1, f"估值日期：{day}")
    for row in sheet.iter_rows():
        label = str(row[0].value or "")
        if label.startswith(("基金单位净值", "累计单位净值", "昨日单位净值")):
            row[1].value = float(nav)
        elif label.startswith("净值日增长率"):
            row[1].value = 0
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _import_dates(client, engine, dates):
    batch_id = client.post("/api/v1/imports", json={}).json()["data"]["id"]
    for day, nav in dates:
        uploaded = client.post(
            f"/api/v1/imports/{batch_id}/files",
            files={"file": (f"{day}.xlsx", _history_workbook(day, nav))},
        )
        assert uploaded.status_code == 201
    assert client.post(f"/api/v1/imports/{batch_id}/complete").status_code == 200
    with Session(engine) as session:
        result = process_next_job(session, client.app.state.settings)
        assert result[0].status == JobStatus.SUCCEEDED
        assert result[1].published_files == len(dates)
        return result[1]


def _run_analysis(client, engine):
    with Session(engine) as session:
        result = process_next_job(session, client.app.state.settings)
        assert result[0].status == JobStatus.SUCCEEDED


@pytest.mark.parametrize(
    "corrections",
    [
        [("2026-08-26", "1.1")],
        [("2026-08-25", "1.1")],
        [("2026-08-25", "1.05"), ("2026-08-26", "1.1")],
    ],
)
def test_historical_auto_import_invalidates_and_rebuilds_later_metrics(
    admin_client, app_and_engine, corrections
):
    engine = app_and_engine[1]
    with Session(engine) as session:
        fund = Fund(standard_name="千金一号")
        session.add(fund)
        session.commit()
        fund_id = fund.id
    _import_dates(admin_client, engine, [("2026-08-25", "1"), ("2026-08-27", "1.21")])
    _run_analysis(admin_client, engine)
    url = f"/api/v1/funds/{fund_id}/nav-series"
    baseline = admin_client.get(url).json()
    assert Decimal(baseline["data"]["points"][-1]["daily_return"]) == Decimal("0.21")

    imported = _import_dates(admin_client, engine, corrections)
    pending = admin_client.get(url).json()
    assert pending["data"]["points"][-1]["analysis_status"] == "pending"
    assert (
        admin_client.get("/api/v1/dashboard/overview").json()["meta"]["analysis_status"]
        == "pending"
    )
    _run_analysis(admin_client, engine)

    result = admin_client.get(url).json()
    last = result["data"]["points"][-1]
    assert last["valuation_date"] == date(2026, 8, 27).isoformat()
    assert Decimal(last["daily_return"]) == Decimal("0.1")
    assert last["metric_source"] == "persisted"
    assert last["analysis_run_id"] == imported.analysis_run_ids[0]
    assert result["meta"]["analysis_status"] == "ready"
    overview = admin_client.get("/api/v1/dashboard/overview").json()
    assert Decimal(overview["data"]["company_daily_return"]) == Decimal("0.1")
    assert overview["meta"]["analysis_run_id"] == imported.analysis_run_ids[0]
